import openpyxl


def export_excel(table, value, column, row):
    workbook = openpyxl.load_workbook(table)
    sheet = workbook.active
    if type(value) == list:
        value = ', '.join(value)
    sheet[row][column].value = value
    workbook.save(table)


def quantity_row(table):
    workbook = openpyxl.open(table, read_only=True)
    sheet = workbook.active

    return sheet.max_row + 1


def repeat_del(table):
    workbook = openpyxl.load_workbook(table)
    sheet = workbook.active

    # values_list = []
    # for cell in sheet['L']:
    #     values_list.append(cell.value)

    values_list = [cell.value for cell in sheet['L']]

    row_num = 1
    for row in sheet.iter_rows(min_row=2, min_col=12, max_col=12, values_only=True):
        row_num += 1
        if values_list.count(row[0]) > 1:
            sheet.delete_rows(row_num, 1)
    workbook.save(table)


if __name__ == "__main__":
    repeat_del('Таблица прицепы.xlsx')
