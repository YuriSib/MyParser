from urllib.parse import quote
import requests
from bs4 import BeautifulSoup
from fuzzywuzzy import fuzz
from collections import Counter
import openpyxl


def compare_similarity(text, text_list, threshold=60):
    similar_texts = []
    idx = 0
    for item in text_list:
        similarity_ratio = fuzz.ratio(text.lower(), item.lower())
        if similarity_ratio > threshold:
            similar_texts.append((item, idx))
        idx += 1
    return similar_texts


def search_xml(qwery):
    url = f'https://xmlstock.com/yandex/xml/?user=11362&key=2a81ea2bf46144411cc5e8c148f5fcfa&query={quote(qwery)}' \
          f'&groupby=attr%3D%22%22.mode%3Dflat.groups-on-page%3D30.docs-in-group%3D1'
    response = requests.get(url)

    dirty_list_link = BeautifulSoup(response.text, features="xml").find_all('url')
    list_link = [link.text.strip('</url>') for link in dirty_list_link]

    dirty_list_name = BeautifulSoup(response.text, features="xml").find_all('title')
    list_name = [name.get_text() for name in dirty_list_name]

    filtered_list_name = compare_similarity(qwery, list_name)

    beginner_filtered_list_link = [list_link[name[1]] for name in filtered_list_name]
    filtered_list_link = [link.split('/')[2] for link in beginner_filtered_list_link]
    finished_list_link = list(Counter(filtered_list_link).keys())

    return finished_list_link


def import_xl(table):
    catalog = openpyxl.open(table)
    sheet = catalog.active
    cnt = 1
    for row in range(1, (sheet.max_row + 1)):
        article = sheet[row][1].value
        name = sheet[row][0].value
        print(cnt)
        cnt += 1

        if article:
            response = str(article) + ' ' + name
            links = ' '.join(search_xml(response))
            sheet.cell(row=row, column=3).value = links if links else 'не найдено'
        else:
            links = ' '.join(search_xml(name))
            sheet.cell(row=row, column=3).value = links if links else 'не найдено'

        catalog.save(table)


def count_link(table):
    catalog = openpyxl.open(table)
    sheet = catalog.active
    cnt = 1

    list_links = []
    for row in range(1, 2020):
        print(cnt)
        cnt += 1
        str_links = sheet[row][2].value
        links = str_links.split(' ')
        # links = [link for link in str_links]
        list_links.extend(links)

    counts = Counter(list_links)

    with open('output.txt', 'w', encoding='utf-8') as file:
        for value, count in counts.items():
            if count > 150:
                file.write(f'{value} повторяется {count} раз(а)\n')


if __name__ == '__main__':
    count_link('Для анализа канцтоваров.xlsx')


