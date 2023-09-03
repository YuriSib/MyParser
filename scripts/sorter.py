import openpyxl


def sort_excel_by_second_column(table_path):
    workbook = openpyxl.load_workbook(table_path)
    sheet = workbook.active
    data = sheet.iter_rows(values_only=True)

    sorted_data = sorted(data, key=lambda x: x[3] if x[3] is not None else '')  # Сортировка по второй колонке

    new_rows = []
    for row in sorted_data:
        new_rows.append(row)

    # Очищаем старую таблицу от данных
    sheet.delete_rows(1, sheet.max_row)

    # Записываем отсортированные данные в старую таблицу
    for row_idx, row in enumerate(new_rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    workbook.save(table_path)


file_path = 'Электроинструмент.xlsx'  # Укажите путь к вашему Excel файлу
sort_excel_by_second_column(file_path)
